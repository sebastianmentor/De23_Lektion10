from openpyxl import Workbook
from openpyxl import load_workbook

# Create a new Excel workbook and select the active sheet
workbook = Workbook()
sheet = workbook.active
sheet.title = "My Sheet"

# Write some data to the sheet
sheet['A1'] = 'Hello'
sheet['B1'] = 'World!'
sheet['A2'] = 'This is'
sheet['B2'] = 'openpyxl!'

# Save the workbook to a file
filename = 'my_workbook.xlsx'
workbook.save(filename)

# Load the workbook we just created
loaded_workbook = load_workbook(filename)
loaded_sheet = loaded_workbook['My Sheet']

# Read data back from the sheet
print("Reading data from the Excel file:")
for row in loaded_sheet.iter_rows(min_row=1, max_col=2, max_row=2, values_only=True):
    print(row)
