from openpyxl import Workbook

wb = Workbook()

ws = wb.active
print(ws)

ws1 = wb.create_sheet("Mysheet") # insert at the end (default)
# or
# ws2 = wb.create_sheet("Mysheet", 0) # insert at first position
# or
# ws3 = wb.create_sheet("Mysheet", -1) # insert at the penultimate position

ws.title = "New Title"
ws3 = wb["New Title"]
print(wb.sheetnames)